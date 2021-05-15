from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook

TOKEN = '1656164530:AAF-4RoNDN6ei2TEhFHbagbJ_FNOAL26f9Y'
book = load_workbook('база данных.xlsx')
sheet_1 = book['Лист1']
schedule_page = book['расписание']


def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения из телеграм
    dispatcher = updater.dispatcher
    handler = MessageHandler(Filters.all, do_echo)  # отфильтровываем сообщения: теперь устройство должно реагировать
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    schedule_handler = MessageHandler(Filters.all, do_schedule)

    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(schedule_handler)
    dispatcher.add_handler(handler)
    updater.start_polling()
    updater.idle()


def do_echo(update, context):
    text = update.message.text

    if text == "привет":
        update.message.reply_text(text="доброго времени суток, друг")
    else:
        update.message.reply_text(text="Привет! Нажми /start чтобы запустить меня:)")



def do_start(update, context):
    keyboard = [
        ["понедельник", "вторник", "среда", "четверг", "пятница"]
    ]
    update.message.reply_text(
        text="Выбери день чтобы узнать расписание",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))


def do_help(update, context):
    update.message.reply_text(text="что случилось? Всё ж было норм")



# def do_sticker(update: Update, context):
#     sticker_id = update.message.sticker.file_id
#     update.message.reply_text(sticker_id)
#     update.message.reply_sticker(sticker_id)


def get_schedule(day):
    reply_text = ""
    for i in range(1, 9):
        #  print(schedule_page.cell(row=i, column=1).value)
        reply_text += schedule_page.cell(row=i, column=day).value + "\n"
    return reply_text

def do_schedule(update, context):
    text = update.message.text
    days = ["понедельник", "вторник", "среда", "четверг", "пятница"]
    if text == "привет":
        update.message.reply_text(text="Привет! Нажми /start чтобы запустить меня:)")
        return

    for i in days:
        if text == i:
            if text == "понедельник":
                update.message.reply_text(text=get_schedule(1))
            elif text == "вторник":
                update.message.reply_text(text=get_schedule(2))
            elif text == "среда":
                update.message.reply_text(text=get_schedule(3))
            elif text == "четверг":
                update.message.reply_text(text=get_schedule(4))
            elif text == "пятница":
                update.message.reply_text(text=get_schedule(5))
            return


    update.message.reply_text(text="Ты втираешь мне какую-то дичь. Нажми /start чтобы запустить меня:)")





    # print(stickers_page.max_row)
    # for row in range(2, stickers_page.max_row + 1):
    #     catch_phrase = stickers_page.cell(row=row, column=4).value
    #     print(catch_phrase)
    #     print(text)
    #     if catch_phrase in text:
    #         sticker_id = stickers_page.cell(row=row, column=3).value
    #         update.message.reply_sticker(sticker_id)





main()