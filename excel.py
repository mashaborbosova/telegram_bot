from openpyxl import load_workbook


book = load_workbook('база данных.xlsx')
sheet_1 = book['Лист1']
schedule_page = book['расписание']

print(book.worksheets)
for i in range(1, 37):
    print(schedule_page.cell(row=i, column=1).value)
